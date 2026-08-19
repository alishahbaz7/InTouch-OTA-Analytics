"""The sharing surface: the digest on every page, and the bundle round trip over HTTP."""

from __future__ import annotations

import io
from datetime import datetime, timedelta

import pytest

pytest.importorskip("fastapi")
from fastapi.testclient import TestClient  # noqa: E402

from ota_analytics import bundle, config, db, identity, ingest, rollup, scheduler, sources  # noqa: E402

from tests.test_pages import PAGES, api_device  # noqa: E402


@pytest.fixture
def client(tmp_path, monkeypatch):
    monkeypatch.setattr(config, "DATA_DIR", tmp_path)
    monkeypatch.setattr(config, "DB_PATH", tmp_path / "test.db")
    monkeypatch.setattr(config, "EXPORT_DIR", tmp_path / "exports")
    monkeypatch.setattr(scheduler, "STATE_PATH", tmp_path / "scheduler.json")
    monkeypatch.setattr(sources, "SETTINGS_PATH", tmp_path / "connection.json")
    monkeypatch.setattr(scheduler, "_scheduler", None)

    conn = db.connect()
    for hours, firmware in ((6, "7.5.0.27"), (0, "7.5.0.51A")):
        result = ingest.ingest_records(
            conn, [api_device("111", firmware, target="7.5.0.51A"),
                   api_device("222", "7.5.0.51A")],
            source_name="API test", snapshot_at=datetime.now() - timedelta(hours=hours))
        rollup.rollup_snapshot(conn, result.snapshot_id)

    from ota_analytics import api
    return TestClient(api.app, raise_server_exceptions=False)


@pytest.mark.parametrize("path", PAGES)
def test_the_digest_appears_on_every_page(client, path):
    """The first question in any disagreement, answerable without navigating anywhere."""
    conn = db.connect()
    expected = identity.short(identity.fleet_digest(conn))
    response = client.get(path)
    assert response.status_code == 200
    assert expected in response.text, f"{path} does not show the fleet digest"


def test_identity_endpoint_reports_both_clocks(client):
    data = client.get("/api/identity").json()
    assert data["snapshots"] == 2
    assert data["digest_short"] == identity.short(data["fleet_digest"])
    # Deliberately separate fields: when the data was true vs when we pulled it.
    assert data["last_snapshot_at"] and data["last_ingest_at"]
    assert data["db_id"] and data["instance_label"]


def test_bundle_download_is_a_readable_bundle(client):
    response = client.get("/update/bundle")
    assert response.status_code == 200
    assert response.content[:4] == bundle.ZIP_MAGIC
    assert bundle.SUFFIX in response.headers["content-disposition"]

    described = bundle.describe(response.content)
    assert described["snapshot_count"] == 2
    assert described["source"]["fleet_digest"] == identity.fleet_digest(db.connect())


def test_bundle_upload_round_trips_through_the_page(client, tmp_path):
    """Download from this install, import it back: idempotent, and it says so."""
    payload = client.get("/update/bundle").content
    response = client.post("/update/bundle-import",
                           files={"file": ("share.otabundle", payload, "application/zip")})
    assert response.status_code == 200
    assert "already loaded" in response.text
    assert db.connect().execute("SELECT COUNT(*) FROM snapshot").fetchone()[0] == 2


def test_a_long_import_does_not_freeze_the_dashboard(client, monkeypatch):
    """The upload routes are the only `async` handlers, so they run on the event loop.

    Calling a job that takes minutes directly from one stops the server answering *anything* —
    not just the import, but every page and /healthz too. The symptom is a browser that spins
    for ever, which reads as "the import hung" when the whole dashboard had.
    """
    import asyncio

    from ota_analytics import bundle as bundle_module

    ran_on_loop = []

    def fake_import(conn, content, **kwargs):
        try:
            asyncio.get_running_loop()
        except RuntimeError:
            pass                      # no loop in this thread: correctly off the event loop
        else:
            ran_on_loop.append(True)
        return bundle_module.ImportResult(status="empty", message="checked")

    monkeypatch.setattr(bundle_module, "import_bundle", fake_import)
    response = client.post("/update/bundle-import",
                           files={"file": ("x.otabundle", b"PK\x03\x04stub", "application/zip")})

    assert response.status_code == 200
    assert not ran_on_loop, "the merge ran on the event loop and would freeze every page"


def test_a_long_ingest_does_not_freeze_the_dashboard(client, monkeypatch):
    """Same hazard on the .xlsx upload route, which also has to be async to read the file."""
    import asyncio

    from ota_analytics import api

    ran_on_loop = []

    def fake_store_and_ingest(filename, content):
        try:
            asyncio.get_running_loop()
        except RuntimeError:
            pass
        else:
            ran_on_loop.append(True)
        return {"level": "ok", "message": "checked"}

    monkeypatch.setattr(api, "_store_and_ingest", fake_store_and_ingest)
    response = client.post("/update/import",
                           files={"file": ("Devices_1_15Aug26_1511.xlsx", b"PK\x03\x04", "x")})

    assert response.status_code == 200
    assert not ran_on_loop, "the ingest ran on the event loop and would freeze every page"


def wait_for_job(client, timeout: float = 20.0) -> dict:
    """Poll until the background job stops running. Returns its final state."""
    import time

    deadline = time.monotonic() + timeout
    while time.monotonic() < deadline:
        state = client.get("/api/progress").json()
        if not state.get("active") or state.get("status") != "running":
            return state
        time.sleep(0.05)
    raise AssertionError("the job never finished")


def test_uploading_something_that_is_not_a_bundle_is_reported_not_raised(client):
    """The POST now returns straight away, so the failure arrives through the job, not the
    response body — but it must still arrive, and still be readable."""
    from ota_analytics import progress

    progress.clear()
    response = client.post("/update/bundle-import", follow_redirects=False,
                           files={"file": ("login.html", b"<html>Sign in</html>", "text/html")})
    assert response.status_code == 303          # answered immediately, not after the work

    state = wait_for_job(client)
    assert state["status"] == "error"
    assert "not a bundle" in state["message"]
    assert "Something went wrong" not in client.get("/update").text
    progress.clear()


def test_renaming_the_install_shows_up_in_the_footer(client):
    response = client.post("/update/label", data={"label": "shahbaz-laptop"})
    assert response.status_code == 200
    assert "shahbaz-laptop" in client.get("/").text


def test_xlsx_export_carries_its_provenance(client):
    """A report that has been emailed around still says which dataset produced it."""
    from openpyxl import load_workbook

    response = client.get("/devices/export?format=xlsx")
    assert response.status_code == 200

    book = load_workbook(io.BytesIO(response.content), read_only=True)
    assert "Source" in book.sheetnames
    values = {row[0]: row[1] for row in book["Source"].iter_rows(values_only=True) if row[0]}
    assert values["Fleet digest"] == identity.short(identity.fleet_digest(db.connect()))
    assert values["Snapshots held"] == "2"


def test_imei_list_export_is_not_stamped(client):
    """The IMEI list is pasted straight into the platform — nothing may be added to it."""
    response = client.get("/devices/export?format=txt")
    assert response.status_code == 200
    for line in response.text.splitlines():
        assert line.isdigit(), f"non-IMEI line in the paste list: {line!r}"
