"""Downloads. The file has to match the table it was taken from, exactly.

An export that silently returns a different set than the page shows is worse than no export at
all, because someone will act on it.
"""

from __future__ import annotations

import csv
import io

import pytest

pytest.importorskip("fastapi")

from ota_analytics import exports  # noqa: E402
from tests.test_pages import client  # noqa: F401,E402  (reuses the seeded dashboard)


def rows_from_csv(body: str) -> list[dict]:
    return list(csv.DictReader(io.StringIO(body.lstrip("﻿"))))


# ─── characters a spreadsheet refuses ───────────────────────────────────────

# Exactly the shape found in the real export: 128 devices carry an ICCID that is a valid number
# followed by a backspace and stray bytes. openpyxl raises IllegalCharacterError rather than
# writing it, so one bad cell failed the whole download with a 500 — and only the .xlsx option,
# which is why it looked like that format was broken rather than the data.
CORRUPT_ICCID = "8991922406995209166F\x08\x08Áá"


def test_xlsx_survives_a_control_character_in_the_data():
    payload = exports.to_xlsx([{"imei": "111", "iccid": CORRUPT_ICCID}],
                              [("imei", "IMEI"), ("iccid", "ICCID")])
    assert payload[:4] == b"PK\x03\x04"

    from openpyxl import load_workbook
    sheet = load_workbook(io.BytesIO(payload), read_only=True).active
    values = [list(row) for row in sheet.iter_rows(values_only=True)]
    # The number survives; only the illegal bytes are dropped.
    assert values[1][1] == "8991922406995209166FÁá"


def test_csv_strips_them_too():
    """A raw backspace in a CSV is just as broken — it fails later, in whatever opens it."""
    body = exports.to_csv([{"imei": "111", "iccid": CORRUPT_ICCID}],
                          [("imei", "IMEI"), ("iccid", "ICCID")])
    assert "\x08" not in body
    assert "8991922406995209166F" in body


def test_legal_whitespace_is_left_alone():
    """Tab, newline and carriage return are valid in both formats and carry meaning."""
    assert exports._clean("a\tb\nc\rd") == "a\tb\nc\rd"


def test_the_whole_device_export_survives_corrupt_data(client):  # noqa: F811
    """End to end through the route, because the failure was a 500 on the real fleet."""
    response = client.get("/devices/export?format=xlsx")
    assert response.status_code == 200
    assert response.content[:4] == b"PK\x03\x04"


# ─── formatting ─────────────────────────────────────────────────────────────

def test_csv_has_a_header_and_one_row_per_device():
    rows = [{"imei": "111", "firmware": "7.5.0.51A"}, {"imei": "222", "firmware": "7.5.0.27"}]
    body = exports.to_csv(rows, [("imei", "IMEI"), ("firmware", "Firmware")])
    parsed = rows_from_csv(body)

    assert [r["IMEI"] for r in parsed] == ["111", "222"]
    assert parsed[0]["Firmware"] == "7.5.0.51A"


def test_missing_values_become_blank_not_the_word_none():
    body = exports.to_csv([{"imei": "111", "firmware": None}],
                          [("imei", "IMEI"), ("firmware", "Firmware")])
    assert "None" not in body
    assert rows_from_csv(body)[0]["Firmware"] == ""


def test_imei_list_is_one_per_line_and_nothing_else():
    """This is pasted straight into the platform, so any extra character breaks it."""
    text = exports.to_imei_list([{"imei": "865510082004294"}, {"imei": "869742089244468"}])
    assert text == "865510082004294\n869742089244468\n"


def test_imei_list_skips_rows_without_one():
    assert exports.to_imei_list([{"imei": "111"}, {"imei": None}, {}]) == "111\n"


def test_xlsx_is_a_real_workbook_with_a_frozen_header():
    from openpyxl import load_workbook

    payload = exports.to_xlsx([{"imei": "111", "firmware": "7.5.0.51A"}],
                              [("imei", "IMEI"), ("firmware", "Firmware")], "Devices")
    sheet = load_workbook(io.BytesIO(payload)).active

    assert sheet.title == "Devices"
    assert [c.value for c in sheet[1]] == ["IMEI", "Firmware"]
    assert sheet["A2"].value == "111"
    assert sheet.freeze_panes == "A2"


def test_imei_column_is_text_so_excel_does_not_mangle_it():
    """865510082004294 as a number becomes 8.65510E+14 — and then it is useless."""
    from openpyxl import load_workbook

    payload = exports.to_xlsx([{"imei": "865510082004294"}], [("imei", "IMEI")])
    sheet = load_workbook(io.BytesIO(payload)).active
    assert sheet["A2"].number_format == "@"


def test_filenames_carry_a_timestamp():
    name = exports.timestamped("devices", "csv")
    assert name.startswith("devices_") and name.endswith(".csv")


# ─── the routes ─────────────────────────────────────────────────────────────

@pytest.mark.parametrize("path", ["/devices/export", "/changes/export", "/pending/export"])
@pytest.mark.parametrize("fmt", ["csv", "xlsx", "txt"])
def test_every_export_downloads(client, path, fmt):  # noqa: F811
    response = client.get(f"{path}?format={fmt}")
    assert response.status_code == 200
    assert "attachment; filename=" in response.headers["content-disposition"]
    assert response.content


def test_a_device_export_matches_the_filtered_page(client):  # noqa: F811
    """Same filter, same rows — the property that makes the download safe to act on."""
    page = client.get("/devices?status=Online")
    body = client.get("/devices/export?status=Online&format=csv").text
    rows = rows_from_csv(body)

    for row in rows:
        assert row["Status"] == "Online"
    # every exported IMEI appears in the page it claims to mirror
    for row in rows:
        assert row["IMEI"] in page.text


def test_an_export_is_not_limited_to_one_page(client):  # noqa: F811
    """The page shows 100 rows; the file must hold everything that matched."""
    all_rows = rows_from_csv(client.get("/devices/export?format=csv").text)
    page_rows = rows_from_csv(client.get("/devices/export?format=csv&page=1").text)
    assert len(all_rows) == len(page_rows) >= 4


def test_fallback_only_export_contains_just_the_fallbacks(client):  # noqa: F811
    body = client.get("/changes/export?window=all&only=fallbacks&format=csv").text
    rows = rows_from_csv(body)
    assert rows, "the seeded database contains one fallback"
    for row in rows:
        assert row["Verdict"] == "fallback to base"
        assert row["To"] == row["Base"]


def test_change_export_spells_out_the_verdict(client):  # noqa: F811
    rows = rows_from_csv(client.get("/changes/export?window=all&format=csv").text)
    verdicts = {r["Verdict"] for r in rows}
    assert verdicts <= {"upgrade", "planned rollback", "unplanned", "fallback to base"}


def test_imei_download_is_pasteable(client):  # noqa: F811
    text = client.get("/devices/export?format=txt").text
    lines = [line for line in text.splitlines() if line]
    assert lines
    assert all(line.strip() == line and " " not in line for line in lines)


def test_the_imei_file_has_no_byte_order_mark(client):  # noqa: F811
    """A BOM is right for CSV and wrong here — it would corrupt the first IMEI on paste."""
    body = client.get("/devices/export?format=txt").content
    assert not body.startswith(b"\xef\xbb\xbf")
    assert body.split(b"\n")[0].isdigit()


def test_csv_keeps_its_byte_order_mark_for_excel(client):  # noqa: F811
    body = client.get("/devices/export?format=csv").content
    assert body.startswith(b"\xef\xbb\xbf")


def test_an_empty_result_still_downloads_cleanly(client):  # noqa: F811
    """A filter matching nothing should give an empty file, not an error."""
    response = client.get("/devices/export?firmware=does-not-exist&format=csv")
    assert response.status_code == 200
    assert rows_from_csv(response.text) == []
