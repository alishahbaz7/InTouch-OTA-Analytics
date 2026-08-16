"""Downloadable views of whatever is on screen.

Three formats, because they serve different jobs:

    csv   opens anywhere, good for sharing and for pivoting
    xlsx  keeps column widths and a frozen header for someone reading it directly
    txt   just the IMEIs, one per line — the format you paste back into the OTA platform
          to act on a cohort

Whatever is exported matches the filters in force on the page. An export that quietly returned
something else would be worse than none, because the action taken from it would be wrong.
"""

from __future__ import annotations

import csv
import io
from datetime import datetime

# (key in the row dict, column header)
DEVICE_COLUMNS = [
    ("imei", "IMEI"),
    ("device_model", "Model"),
    ("firmware", "Firmware"),
    ("fallback_tag", "Fallback"),
    ("prev_firmware", "Previous firmware"),
    ("update_firmware", "Target firmware"),
    ("base_firmware", "Base firmware"),
    ("configuration", "Configuration"),
    ("hw_ver", "Hardware"),
    ("status", "Status"),
    ("queue_state", "Task state"),
    ("queue", "Pending tasks"),
    ("seen_at", "Last seen"),
    ("seen_age_hours", "Hours since seen"),
    ("last_fw_change_at", "Last firmware change"),
    ("last_checked_at", "Last checked"),
    ("groups_raw", "Groups"),
    ("vin", "VIN"),
    ("iccid", "ICCID"),
]

CHANGE_COLUMNS = [
    ("changed_at", "Changed at"),
    ("imei", "IMEI"),
    ("device_model", "Model"),
    ("from_firmware", "From"),
    ("to_firmware", "To"),
    ("direction", "Direction"),
    ("verdict", "Verdict"),
    ("update_firmware", "Target"),
    ("base_firmware", "Base"),
    ("hw_ver", "Hardware"),
    ("status", "Status"),
    ("queue_state", "Task state"),
    ("groups_raw", "Groups"),
]


def timestamped(stem: str, suffix: str) -> str:
    return f"{stem}_{datetime.now().strftime('%d%b%y_%H%M')}.{suffix}"


def describe(filters: dict) -> str:
    """A short slug describing the filters, so a downloaded file says what it holds."""
    parts = [str(value).replace(" ", "-") for value in filters.values() if value]
    return "_".join(parts)[:60]


def to_csv(rows: list[dict], columns) -> str:
    buffer = io.StringIO()
    writer = csv.writer(buffer, lineterminator="\n")
    writer.writerow([header for _, header in columns])
    for row in rows:
        writer.writerow([_clean(row.get(key)) for key, _ in columns])
    return buffer.getvalue()


def to_imei_list(rows: list[dict]) -> str:
    """One IMEI per line — ready to paste into the platform's device selector."""
    return "\n".join(str(row["imei"]) for row in rows if row.get("imei")) + "\n"


# Order matters: this is read top to bottom by someone comparing two files.
PROVENANCE_FIELDS = [
    ("instance_label", "Produced by"),
    ("digest_short", "Fleet digest"),
    ("snapshots", "Snapshots held"),
    ("first_snapshot_at", "Coverage from"),
    ("last_snapshot_at", "Coverage to"),
    ("last_ingest_at", "Last fetched from platform"),
    ("db_id", "Database id"),
    ("created_at", "Database created"),
    ("fleet_digest", "Fleet digest (full)"),
]


def _write_provenance(workbook, provenance: dict) -> None:
    """Record which dataset produced the file, on its own sheet.

    On a sheet rather than above the data because a report is opened, filtered and pivoted —
    header rows in the way break every one of those. Two people holding the same numbers can
    still be looking at different snapshot sets, and once a file leaves the dashboard nothing
    else says which one it came from.
    """
    sheet = workbook.create_sheet("Source")
    sheet.append(["Field", "Value"])
    sheet["A1"].font = sheet["B1"].font = _bold()
    for key, label in PROVENANCE_FIELDS:
        value = provenance.get(key)
        sheet.append([label, "" if value is None else str(value)])
    sheet.append(["Report generated", datetime.now().strftime("%d %b %Y %H:%M")])
    sheet.append([])
    sheet.append(["Two reports agree only if the fleet digest above matches. It is a "
                  "fingerprint of the exports loaded, not of the file."])
    sheet.column_dimensions["A"].width = 30
    sheet.column_dimensions["B"].width = 46


def _bold():
    from openpyxl.styles import Font
    return Font(bold=True)


def to_xlsx(rows: list[dict], columns, sheet_name: str = "Devices",
            provenance: dict | None = None) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    workbook = Workbook(write_only=False)
    sheet = workbook.active
    sheet.title = sheet_name[:31]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2A313B")
    sheet.append([header for _, header in columns])
    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")

    for row in rows:
        sheet.append([_clean(row.get(key)) for key, _ in columns])

    # IMEIs are identifiers, not numbers: Excel would render 865510082004294 in scientific
    # notation and a copy-paste back into the platform would then be wrong.
    for cell in sheet["A"][1:]:
        cell.number_format = "@"

    for index, (_, header) in enumerate(columns, start=1):
        width = max(len(header) + 2, 12)
        if header in {"Groups", "Last seen", "Last firmware change", "Last checked"}:
            width = 22
        sheet.column_dimensions[get_column_letter(index)].width = width

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    if provenance:
        _write_provenance(workbook, provenance)

    stream = io.BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def _clean(value):
    """Excel and CSV both prefer plain scalars; round the one float we carry."""
    if value is None:
        return ""
    if isinstance(value, float):
        return round(value, 1)
    return value
