"""Ingest a platform export into the snapshot warehouse.

Append-only and idempotent: a file is identified by the SHA-256 of its bytes, so re-ingesting
the same export is a no-op rather than a duplicate. Reads are streamed because the exports run
to tens of megabytes.

**Two formats, one path.** `.xlsx` is what the platform produces; `.csv` is accepted because a
colleague's data does not always arrive as a spreadsheet. Both go through the same header
mapping and the same normalization, so a CSV cannot take a shortcut past a rule the spreadsheet
obeys. Columns are matched by name, not position, so column order does not matter either way.

A CSV holding a report *this dashboard* produced is loadable but second-hand, and ingest says
so: its values have already been through normalization once, so `device_model_raw` holds a
canonical name rather than the platform's original spelling, and columns the report does not
carry (First Ping, the raw VIN) are absent rather than empty. A quality finding records that,
because it is not visible from the numbers afterwards.
"""

from __future__ import annotations

import csv
import hashlib
import re
import sqlite3
import time
from contextlib import contextmanager
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Iterator

from openpyxl import load_workbook

# Formats ingest can read. Order matters only for reporting.
EXPORT_SUFFIXES = (".xlsx", ".csv")

from . import config, normalize, quality, registry

# Source field -> our column name, keyed by a flattened form of the name (lowercase, letters and
# digits only). One map serves both sources: the spreadsheet's "Device Name/VIN" and an API's
# "deviceNameVin" both flatten to "devicenamevin", so neither breaks when the platform renames,
# reorders or re-cases things.
FIELD_ALIASES = {
    # Spreadsheet header    |  API field (verified against /fotaAdminApi/api/user/devices)
    "imei": "imei",            "deviceid": "imei", "deviceimei": "imei",
    "status": "status", "devicestatus": "status", "connectionstatus": "status",
    "queue": "queue", "queuecount": "queue", "pendingtasks": "queue", "pendingcount": "queue",
    "typetask": "queue",       # {} = done, {"1": n} = pending, key absent = never tasked
    "devicenamevin": "device_name", "devicename": "device_name",
    "createdby": "created_by", "createdbyname": "created_by", "creator": "created_by",
    "devicemodel": "device_model_raw", "model": "device_model_raw",
    "modelname": "device_model_raw",
    "firmware": "firmware_raw", "currfirmver": "firmware_raw",
    "firmwareversion": "firmware_raw", "fwversion": "firmware_raw",
    "currentfirmware": "firmware_raw",
    "configuration": "configuration", "currconfigversion": "configuration",
    "configversion": "configuration", "config": "configuration",
    "seenat": "seen_at", "lastpingtime": "seen_at", "lastseen": "seen_at",
    "lastseenat": "seen_at", "lastping": "seen_at", "lastcommunication": "seen_at",
    "iccid": "iccid", "simiccid": "iccid",
    "hwver": "hw_ver", "hwversion": "hw_ver", "hardwareversion": "hw_ver",
    "hardware": "hw_ver",      # heading used by this dashboard's own CSV reports
    "vin": "vin_raw", "vinnumber": "vin_raw",
    "groups": "groups_raw", "groupname": "groups_raw", "groupnames": "groups_raw",
    "firstping": "first_ping", "firstseen": "first_ping", "commissionedat": "first_ping",
    # Rollout intent — present in the API only, dropped by the spreadsheet export.
    "updatefirmver": "update_firmware", "targetfirmware": "update_firmware",
    "basefirm": "base_firmware", "basefirmware": "base_firmware",
    "newconfigversion": "target_config",
    "baseconfig": "base_config",
}

# Fields whose value is an epoch number rather than a formatted date.
EPOCH_FIELDS = {"lastpingtime", "firstping", "creationtime", "keyexpiretime"}

_FLATTEN = re.compile(r"[^a-z0-9]")


def flatten_key(name: object) -> str:
    """Reduce a column header or JSON key to a comparable form: 'Device Name/VIN' -> 'devicenamevin'."""
    return _FLATTEN.sub("", str(name).lower())


def map_field(name: object) -> str | None:
    return FIELD_ALIASES.get(flatten_key(name))

# Without these the snapshot cannot be analyzed at all.
REQUIRED_COLUMNS = {"imei", "status", "device_model_raw", "firmware_raw"}

# The columns that describe a device. Order matters: build_device_row returns them in exactly
# this sequence. seen_age_hours is absent on purpose — it is snapshot_at minus seen_at, so it
# differs on every row of every fetch and would make every device look changed. device_state
# computes it at read time instead.
DEVICE_COLUMNS = (
    "imei", "status", "queue", "queue_state", "device_name", "created_by",
    "device_model_raw", "device_model", "firmware_raw", "firmware", "fw_family", "fw_sortkey",
    "configuration", "config_sortkey", "seen_at", "iccid", "hw_ver", "vin", "vin_raw",
    "groups_raw", "first_ping", "update_firmware", "base_firmware", "target_config",
    "base_config",
)

# Rows are collected here first, then compared against stored state in one statement rather than
# one query per device — 35,000 round trips per fetch would cost more than the writes they save.
CREATE_STAGE = f"""
CREATE TEMP TABLE IF NOT EXISTS stage_device (
  {', '.join(f'{c} TEXT' for c in DEVICE_COLUMNS)},
  PRIMARY KEY (imei)
) WITHOUT ROWID
"""

INSERT_STAGE = (f"INSERT OR IGNORE INTO stage_device ({', '.join(DEVICE_COLUMNS)}) "
                f"VALUES ({','.join('?' * len(DEVICE_COLUMNS))})")

# Columns build_device_row computes rather than reads. They are "provided" whenever the column
# they are derived from is — otherwise a firmware change would be written while the canonical
# `firmware` alongside it was carried forward from the previous fetch, which is worse than
# either behaviour on its own.
DERIVED_FROM = {
    "device_model": "device_model_raw",
    "firmware": "firmware_raw",
    "fw_family": "firmware_raw",
    "fw_sortkey": "firmware_raw",
    "config_sortkey": "configuration",
    "vin": "vin_raw",
    "queue_state": "queue",
}


def provided_columns(mapping) -> set[str]:
    """Which device columns this source actually carries, derived ones included.

    `mapping` is source-field -> column, so its values are only the columns read straight from
    the source. Everything computed from one of those counts as provided too.
    """
    provided = set(mapping.values())
    provided |= {column for column, source in DERIVED_FROM.items() if source in provided}
    # The API sends no status field; build_device_row derives it from last-contact age instead.
    if "seen_at" in provided:
        provided.add("status")
    return provided


def insert_changed_sql(provided: set[str]) -> tuple[str, bool]:
    """Build the statement that stores a fetch, given the columns this source actually sent.

    Returns the SQL and whether it carries columns forward, which decides how many parameters
    it binds — the carry-forward join adds one.

    A field the source never sent is *unknown*, not empty, and the difference matters. The
    platform API carries no group information at all, so treating absence as NULL silently wiped
    the groups of 29,384 devices on the first API fetch and every one after it — and groups are
    one of the few dimensions available for explaining why a set of devices reverted. It also
    made every device look changed at the moment the source switched, writing 35,475 rows to
    record nothing.

    So an unsent column keeps the value already on record, and is excluded from the comparison
    that decides whether anything moved: a value this fetch never reported cannot have changed.
    """
    carried = [c for c in DEVICE_COLUMNS if c not in provided and c != "imei"]
    values = ", ".join(("s." + c) if c in provided or c == "imei" else ("prev." + c)
                       for c in DEVICE_COLUMNS)

    # `IS` rather than `=` so NULL compares equal to NULL — most devices have several null
    # columns, and `=` would call every one of them a change.
    compared = [c for c in DEVICE_COLUMNS if c != "imei" and c in provided]
    unchanged = " AND ".join(f"p.{c} IS s.{c}" for c in compared) or "1"

    join = ""
    if carried:
        # Only needed when something has to be carried forward; otherwise it is pure cost.
        join = ("LEFT JOIN device_state prev "
                "ON prev.snapshot_id = (SELECT MAX(id) FROM snapshot WHERE id < ?) "
                "AND prev.imei = s.imei")

    return f"""
INSERT OR IGNORE INTO device_snapshot (snapshot_id, present, {', '.join(DEVICE_COLUMNS)})
SELECT ?, 1, {values}
FROM stage_device s
{join}
WHERE NOT EXISTS (
  SELECT 1 FROM device_snapshot p
  WHERE p.imei = s.imei
    AND p.snapshot_id = (SELECT MAX(x.snapshot_id) FROM device_snapshot x
                         WHERE x.imei = s.imei AND x.snapshot_id < ?)
    AND p.present = 1
    AND {unchanged}
)
""", bool(carried)

# A device the platform has stopped listing. Without an explicit marker this is indistinguishable
# from "nothing changed", and device_state would keep serving its last known values for ever.
INSERT_TOMBSTONES = """
INSERT OR IGNORE INTO device_snapshot (snapshot_id, imei, present)
SELECT ?, d.imei, 0
FROM device_state d
WHERE d.snapshot_id = ?
  AND NOT EXISTS (SELECT 1 FROM stage_device s WHERE s.imei = d.imei)
"""


class IngestError(Exception):
    """Raised when a file cannot be ingested at all."""


@dataclass
class IngestResult:
    path: Path
    status: str  # ingested | already_ingested
    snapshot_id: int | None = None
    snapshot_at: datetime | None = None
    ts_source: str = "filename"
    rows: int = 0            # devices the fetch reported — the fleet size, not what was stored
    changed_rows: int = 0    # rows actually written: what changed, plus any tombstones
    groups: int = 0
    skipped_no_imei: int = 0
    duplicate_imei: int = 0
    unknown_columns: list[str] = field(default_factory=list)
    duration_ms: int = 0
    findings: list = field(default_factory=list)


def build_device_row(imei: str, record: dict, snapshot_iso: str) -> tuple:
    """Normalize one device into the row shape `DEVICE_COLUMNS` describes.

    Shared by both sources so a spreadsheet row and an API record cannot drift apart in how
    they are cleaned — the normalization rules are the analytical contract.
    """
    firmware = normalize.canon_firmware(record.get("firmware_raw"))
    configuration = normalize.clean(record.get("configuration"))
    seen_at = normalize.parse_dt(record.get("seen_at"))
    seen_age = normalize.hours_between(snapshot_iso, seen_at)
    queue_count, queue_state = normalize.parse_queue(record.get("queue"))

    # The API sends no status field, so derive it from last-contact age using the platform's
    # own 24-hour rule. Spreadsheet rows carry STATUS and keep it.
    status = (normalize.canon_status(record["status"]) if record.get("status") is not None
              else normalize.status_from_age(seen_age))

    return (
        imei,
        status,
        queue_count,
        queue_state,
        normalize.clean(record.get("device_name")),
        normalize.clean(record.get("created_by")),
        normalize.clean(record.get("device_model_raw")),
        normalize.canon_model(record.get("device_model_raw")),
        normalize.clean(record.get("firmware_raw")),
        firmware,
        normalize.fw_family(firmware),
        normalize.fw_sortkey(firmware),
        configuration,
        normalize.fw_sortkey(configuration),
        seen_at,
        normalize.clean(record.get("iccid")),
        normalize.clean(record.get("hw_ver")),
        normalize.canon_vin(record.get("vin_raw")),
        normalize.clean(record.get("vin_raw")),
        normalize.clean(record.get("groups_raw")),
        normalize.parse_dt(record.get("first_ping")),
        normalize.canon_firmware(record.get("update_firmware")),
        normalize.canon_firmware(record.get("base_firmware")),
        normalize.clean(record.get("target_config")),
        normalize.clean(record.get("base_config")),
    )


def _open_stage(cursor: sqlite3.Cursor) -> None:
    """Start a fetch with an empty staging table. TEMP tables outlive a single ingest on a
    long-lived connection — the scheduler's — so it is emptied here rather than trusted."""
    cursor.execute(CREATE_STAGE)
    cursor.execute("DELETE FROM stage_device")


def _store_devices(conn: sqlite3.Connection, cursor: sqlite3.Cursor, snapshot_id: int,
                   provided: set[str]) -> int:
    """Write the staged fetch as change rows plus tombstones. Returns rows actually stored.

    This is where a fetch stops costing one row per device. Everything the fetch reported is in
    stage_device; only what differs from the last stored row for that device is kept. `provided`
    is the set of columns this source actually sent — anything outside it is carried forward
    rather than overwritten with NULL.
    """
    sql, carries = insert_changed_sql(provided)
    # The placeholders read in source order: the snapshot being written, then the carry-forward
    # join's "previous snapshot", then the comparison's "last row before this one".
    params = (snapshot_id, snapshot_id, snapshot_id) if carries else (snapshot_id, snapshot_id)
    cursor.execute(sql, params)
    stored = cursor.rowcount

    previous = conn.execute("SELECT MAX(id) AS id FROM snapshot WHERE id < ?",
                            (snapshot_id,)).fetchone()["id"]
    if previous is not None:
        cursor.execute(INSERT_TOMBSTONES, (snapshot_id, previous))
        stored += cursor.rowcount

    cursor.execute("DELETE FROM stage_device")
    return stored


def sha256_file(path: Path, chunk: int = 1 << 20) -> str:
    digest = hashlib.sha256()
    with open(path, "rb") as fh:
        while block := fh.read(chunk):
            digest.update(block)
    return digest.hexdigest()


def _resolve_snapshot_at(path: Path) -> tuple[datetime, str]:
    """Snapshot time comes from the filename; mtime is a flagged fallback."""
    parsed = normalize.snapshot_at_from_filename(path.name)
    if parsed:
        return parsed, "filename"
    return datetime.fromtimestamp(path.stat().st_mtime), "mtime"


@contextmanager
def _table_rows(path: Path) -> Iterator[Iterator[tuple]]:
    """Stream an export's rows, whichever format it is in.

    Both formats are streamed rather than loaded: the spreadsheets run to 22 MB and the CSVs to
    5 MB, and there is no reason to hold either in memory. Everything downstream sees the same
    shape — a header tuple followed by row tuples — so no rule can apply to one format and not
    the other.
    """
    if path.suffix.lower() == ".csv":
        # utf-8-sig because Excel writes a byte-order mark, and our own CSV export adds one so
        # Excel reads accents correctly. Left in place it would become part of the first header
        # name and stop 'IMEI' matching.
        handle = open(path, "r", encoding="utf-8-sig", newline="")
        try:
            yield (tuple(row) for row in csv.reader(handle))
        finally:
            handle.close()
        return

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        yield workbook[workbook.sheetnames[0]].iter_rows(values_only=True)
    finally:
        workbook.close()


# Headings only this dashboard's own reports carry. They are derived columns, so their presence
# means the file is a report rather than a platform export — worth recording, because a report's
# values have already been normalized once and it drops columns the platform sends.
REPORT_ONLY_HEADERS = {"previousfirmware", "hourssinceseen", "lastfirmwarechange", "lastchecked",
                       "fallback"}


def looks_like_dashboard_report(header_row: tuple) -> bool:
    flattened = {flatten_key(cell) for cell in header_row if cell is not None}
    return len(flattened & REPORT_ONLY_HEADERS) >= 2


def _map_headers(header_row: tuple) -> tuple[dict[int, str], list[str]]:
    """Map column positions by header name. Unknown columns are reported, not fatal."""
    mapping: dict[int, str] = {}
    unknown: list[str] = []
    for idx, raw in enumerate(header_row):
        if raw is None:
            continue
        column = map_field(raw)
        if column:
            mapping[idx] = column
        else:
            unknown.append(str(raw).strip())

    missing = REQUIRED_COLUMNS - set(mapping.values())
    if missing:
        raise IngestError(f"export is missing required column(s): {', '.join(sorted(missing))}")
    return mapping, unknown


def ingest_file(conn: sqlite3.Connection, path: Path) -> IngestResult:
    """Ingest one export. Safe to call repeatedly with the same file."""
    path = Path(path)
    if not path.exists():
        raise IngestError(f"file not found: {path}")

    started = time.perf_counter()
    sha = sha256_file(path)

    existing = conn.execute("SELECT id FROM snapshot WHERE file_sha256 = ?", (sha,)).fetchone()
    if existing:
        return IngestResult(path=path, status="already_ingested", snapshot_id=existing["id"])

    snapshot_at, ts_source = _resolve_snapshot_at(path)
    snapshot_iso = snapshot_at.isoformat(sep=" ", timespec="seconds")

    with _table_rows(path) as rows:
        header = next(rows, None)
        if header is None:
            raise IngestError("export is empty")
        mapping, unknown = _map_headers(header)
        from_report = looks_like_dashboard_report(header)

        cursor = conn.cursor()
        cursor.execute(
            "INSERT INTO snapshot (source_file, file_sha256, snapshot_at, ts_source, "
            "row_count, ingested_at, duration_ms) VALUES (?,?,?,?,0,?,0)",
            (path.name, sha, snapshot_iso, ts_source,
             datetime.now().isoformat(sep=" ", timespec="seconds")),
        )
        snapshot_id = cursor.lastrowid

        result = IngestResult(
            path=path, status="ingested", snapshot_id=snapshot_id,
            snapshot_at=snapshot_at, ts_source=ts_source, unknown_columns=unknown,
        )

        _open_stage(cursor)
        seen_imei: set[str] = set()
        device_batch: list[tuple] = []
        group_batch: list[tuple] = []

        for raw_row in rows:
            if raw_row is None:
                continue
            record = {name: raw_row[idx] if idx < len(raw_row) else None
                      for idx, name in mapping.items()}

            imei = normalize.clean(record.get("imei"))
            if not imei:
                result.skipped_no_imei += 1
                continue
            if imei in seen_imei:
                result.duplicate_imei += 1
                continue
            seen_imei.add(imei)

            device_batch.append(build_device_row(imei, record, snapshot_iso))

            for group_name in normalize.split_groups(record.get("groups_raw")):
                group_batch.append((snapshot_id, imei, group_name))

            if len(device_batch) >= config.BATCH_SIZE:
                cursor.executemany(INSERT_STAGE, device_batch)
                result.rows += len(device_batch)
                device_batch.clear()
            if len(group_batch) >= config.BATCH_SIZE:
                cursor.executemany(
                    "INSERT OR IGNORE INTO device_group (snapshot_id, imei, group_name) "
                    "VALUES (?,?,?)", group_batch)
                result.groups += len(group_batch)
                group_batch.clear()

        if device_batch:
            cursor.executemany(INSERT_STAGE, device_batch)
            result.rows += len(device_batch)
        if group_batch:
            cursor.executemany(
                "INSERT OR IGNORE INTO device_group (snapshot_id, imei, group_name) "
                "VALUES (?,?,?)", group_batch)
            result.groups += len(group_batch)
        result.changed_rows = _store_devices(conn, cursor, snapshot_id, provided_columns(mapping))

    result.duration_ms = int((time.perf_counter() - started) * 1000)
    conn.execute(
        "UPDATE snapshot SET row_count = ?, skipped_rows = ?, duration_ms = ? WHERE id = ?",
        (result.rows, result.skipped_no_imei, result.duration_ms, snapshot_id),
    )

    result.findings = quality.run_rules(
        conn, snapshot_id,
        skipped_no_imei=result.skipped_no_imei,
        duplicate_imei=result.duplicate_imei,
        unknown_columns=result.unknown_columns,
        ts_source=result.ts_source,
        from_report=from_report,
    )
    conn.commit()
    registry.apply_snapshot(conn, snapshot_id)
    return result


def ingest_records(conn: sqlite3.Connection, records: list[dict], *,
                   source_name: str, snapshot_at: datetime | None = None,
                   fingerprint: str | None = None) -> IngestResult:
    """Ingest device records fetched from the platform API.

    Same normalization and the same idempotency rule as the spreadsheet path — the fingerprint
    stands in for the file hash, so pulling twice without the fleet changing does not create a
    second identical snapshot.
    """
    started = time.perf_counter()
    if not records:
        raise IngestError("The API returned no device records.")

    snapshot_at = snapshot_at or datetime.now()
    snapshot_iso = snapshot_at.isoformat(sep=" ", timespec="seconds")

    if fingerprint is None:
        digest = hashlib.sha256()
        for record in records:
            digest.update(repr(sorted((str(k), str(v)) for k, v in record.items())).encode())
        fingerprint = digest.hexdigest()

    existing = conn.execute("SELECT id FROM snapshot WHERE file_sha256 = ?",
                            (fingerprint,)).fetchone()
    if existing:
        return IngestResult(path=Path(source_name), status="already_ingested",
                            snapshot_id=existing["id"])

    # Build the field map from every record, not just the first: the API omits keys it has no
    # value for, so an early record can be missing model, firmware or last-ping entirely and
    # sampling only records[0] would silently drop those columns for the whole snapshot.
    all_keys: dict[str, None] = {}
    for record in records:
        all_keys.update(dict.fromkeys(record))

    mapping = {key: column for key in all_keys
               if (column := map_field(key)) is not None}
    unknown = [str(k) for k in all_keys if map_field(k) is None]
    epoch_keys = {key for key in mapping if flatten_key(key) in EPOCH_FIELDS}

    missing = REQUIRED_COLUMNS - set(mapping.values()) - {"status"}
    if missing:
        raise IngestError(
            "The API response is missing required field(s): " + ", ".join(sorted(missing)) +
            f". Fields received: {', '.join(str(k) for k in list(all_keys)[:15])}")

    cursor = conn.cursor()
    cursor.execute(
        "INSERT INTO snapshot (source_file, file_sha256, snapshot_at, ts_source, "
        "row_count, ingested_at, duration_ms) VALUES (?,?,?,?,0,?,0)",
        (source_name, fingerprint, snapshot_iso, "api",
         datetime.now().isoformat(sep=" ", timespec="seconds")))
    snapshot_id = cursor.lastrowid

    result = IngestResult(path=Path(source_name), status="ingested", snapshot_id=snapshot_id,
                          snapshot_at=snapshot_at, ts_source="api", unknown_columns=unknown)

    _open_stage(cursor)
    seen_imei: set[str] = set()
    device_batch: list[tuple] = []
    group_batch: list[tuple] = []

    for raw in records:
        record = {}
        for key, column in mapping.items():
            value = raw.get(key)
            record[column] = (normalize.parse_epoch(value) if key in epoch_keys else value)

        imei = normalize.clean(record.get("imei"))
        if not imei:
            result.skipped_no_imei += 1
            continue
        if imei in seen_imei:
            result.duplicate_imei += 1
            continue
        seen_imei.add(imei)

        device_batch.append(build_device_row(imei, record, snapshot_iso))
        for group_name in normalize.split_groups(record.get("groups_raw")):
            group_batch.append((snapshot_id, imei, group_name))

        if len(device_batch) >= config.BATCH_SIZE:
            cursor.executemany(INSERT_STAGE, device_batch)
            result.rows += len(device_batch)
            device_batch.clear()

    if device_batch:
        cursor.executemany(INSERT_STAGE, device_batch)
        result.rows += len(device_batch)
    if group_batch:
        cursor.executemany(
            "INSERT OR IGNORE INTO device_group (snapshot_id, imei, group_name) VALUES (?,?,?)",
            group_batch)
        result.groups += len(group_batch)
    result.changed_rows = _store_devices(conn, cursor, snapshot_id, provided_columns(mapping))

    result.duration_ms = int((time.perf_counter() - started) * 1000)
    conn.execute("UPDATE snapshot SET row_count = ?, skipped_rows = ?, duration_ms = ? "
                 "WHERE id = ?",
                 (result.rows, result.skipped_no_imei, result.duration_ms, snapshot_id))

    result.findings = quality.run_rules(
        conn, snapshot_id, skipped_no_imei=result.skipped_no_imei,
        duplicate_imei=result.duplicate_imei, unknown_columns=result.unknown_columns,
        ts_source="filename",   # an API timestamp is exact, so no "guessed time" warning
    )
    conn.commit()
    registry.apply_snapshot(conn, snapshot_id)
    return result


def exports_in(directory: Path) -> list[Path]:
    """Every loadable export in a folder, oldest snapshot first so diffs build in order.

    `~$` files are Excel's lock files for a workbook someone has open — they are not exports and
    openpyxl cannot read them.
    """
    files = [p for p in sorted(directory.iterdir())
             if p.is_file() and p.suffix.lower() in EXPORT_SUFFIXES
             and not p.name.startswith("~$")]
    files.sort(key=lambda p: _resolve_snapshot_at(p)[0])
    return files


def ingest_dir(conn: sqlite3.Connection, directory: Path) -> list[IngestResult]:
    """Ingest every export in a folder, oldest snapshot first so diffs build in order."""
    directory = Path(directory)
    if not directory.exists():
        raise IngestError(f"directory not found: {directory}")

    return [ingest_file(conn, path) for path in exports_in(directory)]
